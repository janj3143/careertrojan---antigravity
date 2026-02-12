"""
Extract multi-word collocations from NAICS and SOC classification Excel files.
Outputs structured gazetteer data grouped by category.
"""
import openpyxl
import re
import json
from collections import defaultdict

def clean_text(text):
    """Clean and normalize text."""
    if not text:
        return ""
    text = str(text).strip()
    # Remove leading/trailing punctuation
    text = re.sub(r'^[\s\-–—,;:]+|[\s\-–—,;:]+$', '', text)
    return text

def extract_multiword_phrases(text):
    """Extract meaningful multi-word phrases from text."""
    if not text or not isinstance(text, str):
        return []
    
    phrases = []
    # Split on common delimiters
    parts = re.split(r'[;,\(\)\[\]\{\}]', text)
    
    for part in parts:
        part = clean_text(part)
        words = part.split()
        if len(words) >= 2 and len(words) <= 8:
            # Filter out purely numeric or very short phrases
            if not re.match(r'^[\d\s\.]+$', part) and len(part) > 4:
                phrases.append(part)
    
    return phrases

def categorize_term(term, source_context=""):
    """Categorize a term into the appropriate gazetteer category."""
    term_lower = term.lower()
    ctx_lower = (source_context or "").lower()
    combined = term_lower + " " + ctx_lower
    
    # OIL_GAS
    oil_gas_keywords = ['oil', 'gas', 'petroleum', 'drilling', 'pipeline', 'refiner', 'crude',
                       'wellhead', 'fracking', 'hydraulic fractur', 'lng', 'natural gas',
                       'offshore', 'subsea', 'petrochemical', 'fuel', 'energy', 'mining',
                       'coal', 'geothermal', 'solar', 'wind energy', 'renewable', 'nuclear',
                       'power plant', 'electric util', 'turbine', 'reservoir']
    if any(kw in combined for kw in oil_gas_keywords):
        return 'OIL_GAS'
    
    # BIOTECH_PHARMA
    bio_keywords = ['pharma', 'biotech', 'clinical', 'medical', 'health', 'hospital',
                   'therapeutic', 'diagnostic', 'surgical', 'dental', 'nursing',
                   'physician', 'doctor', 'patient', 'biolog', 'genetic', 'vaccine',
                   'drug', 'medication', 'laboratory', 'patholog', 'radiolog',
                   'oncolog', 'cardio', 'neurol', 'optic', 'veterinar', 'epidemiolog',
                   'anatomy', 'physiolog', 'immunolog', 'microbiol', 'biochem',
                   'biomedi', 'life science', 'healthcare']
    if any(kw in combined for kw in bio_keywords):
        return 'BIOTECH_PHARMA'
    
    # FINANCIAL_SERVICES
    fin_keywords = ['financ', 'bank', 'insurance', 'invest', 'securit', 'credit',
                   'loan', 'mortgage', 'accounting', 'audit', 'tax', 'fiscal',
                   'treasury', 'asset manage', 'wealth', 'broker', 'underwrite',
                   'actuar', 'hedge fund', 'mutual fund', 'pension', 'fiduciar',
                   'compliance', 'risk manage', 'capital market', 'equity',
                   'bond', 'commodit', 'forex', 'stock', 'portfolio']
    if any(kw in combined for kw in fin_keywords):
        return 'FINANCIAL_SERVICES'
    
    # TECH_SKILL
    tech_keywords = ['software', 'computer', 'data', 'algorithm', 'program', 'code',
                    'developer', 'engineer', 'information technology', 'cyber',
                    'network', 'database', 'cloud', 'machine learn', 'artificial intel',
                    'web develop', 'system admin', 'devops', 'full stack',
                    'front end', 'back end', 'mobile app', 'user interface',
                    'user experience', 'information system', 'digital', 'automation',
                    'robotics', 'telecomm', 'semiconductor', 'microprocessor',
                    'circuit', 'embedded system', 'firmware', 'it support',
                    'tech support', 'help desk', 'sas', 'python', 'java', 'sql']
    if any(kw in combined for kw in tech_keywords):
        return 'TECH_SKILL'
    
    # MANUFACTURING
    mfg_keywords = ['manufactur', 'assembl', 'production', 'factory', 'industrial',
                   'machining', 'welding', 'fabricat', 'forging', 'stamping',
                   'casting', 'molding', 'extrusion', 'cnc', 'quality control',
                   'supply chain', 'warehouse', 'logistics', 'lean manufactur',
                   'six sigma', 'injection mold', 'sheet metal', 'tool and die',
                   'process engineer', 'plant manager', 'material handl',
                   'packaging', 'textile', 'steel', 'aluminum', 'plastic',
                   'rubber', 'ceramic', 'glass', 'lumber', 'wood', 'paper',
                   'printing', 'chemical manufactur', 'auto part', 'aerospace']
    if any(kw in combined for kw in mfg_keywords):
        return 'MANUFACTURING'
    
    # Default categories based on source
    if 'soc' in ctx_lower or 'occupation' in ctx_lower:
        return 'JOB_TITLE'
    
    return 'INDUSTRY_TERM'

def read_naics(filepath):
    """Read NAICS descriptions and extract multi-word industry terms."""
    print(f"Reading NAICS file: {filepath}")
    terms = set()
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Sheet: {sheet_name}, rows: {ws.max_row}")
            
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        # Extract the whole cell as a potential term
                        cleaned = clean_text(cell)
                        words = cleaned.split()
                        if 2 <= len(words) <= 8 and len(cleaned) > 4:
                            if not re.match(r'^[\d\s\.]+$', cleaned):
                                terms.add(cleaned)
                        
                        # Also extract sub-phrases
                        for phrase in extract_multiword_phrases(cell):
                            terms.add(phrase)
        wb.close()
    except Exception as e:
        print(f"  Error reading NAICS: {e}")
    
    print(f"  Extracted {len(terms)} unique NAICS terms")
    return terms

def read_soc_descriptions(filepath):
    """Read SOC structure and descriptions."""
    print(f"Reading SOC descriptions: {filepath}")
    terms = set()
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Sheet: {sheet_name}, rows: {ws.max_row}")
            
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        cleaned = clean_text(cell)
                        words = cleaned.split()
                        if 2 <= len(words) <= 8 and len(cleaned) > 4:
                            if not re.match(r'^[\d\s\.]+$', cleaned):
                                terms.add(cleaned)
                        
                        for phrase in extract_multiword_phrases(cell):
                            terms.add(phrase)
        wb.close()
    except Exception as e:
        print(f"  Error reading SOC descriptions: {e}")
    
    print(f"  Extracted {len(terms)} unique SOC description terms")
    return terms

def read_soc_coding_index(filepath):
    """Read SOC coding index for job titles."""
    print(f"Reading SOC coding index: {filepath}")
    terms = set()
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Sheet: {sheet_name}, rows: {ws.max_row}")
            
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        cleaned = clean_text(cell)
                        words = cleaned.split()
                        if 2 <= len(words) <= 8 and len(cleaned) > 4:
                            if not re.match(r'^[\d\s\.]+$', cleaned):
                                terms.add(cleaned)
                        
                        for phrase in extract_multiword_phrases(cell):
                            terms.add(phrase)
        wb.close()
    except Exception as e:
        print(f"  Error reading SOC coding index: {e}")
    
    print(f"  Extracted {len(terms)} unique SOC coding index terms")
    return terms

def main():
    base = r"L:\antigravity_version_ai_data_final\automated_parser"
    
    naics_file = f"{base}\\2022_NAICS_Descriptions.xlsx"
    soc_desc_file = f"{base}\\extendedsoc2020structureanddescriptionsexcel03122025.xlsx"
    soc_index_file = f"{base}\\soc2020volume2thecodingindexexcel03122025.xlsx"
    
    # Read all sources
    naics_terms = read_naics(naics_file)
    soc_desc_terms = read_soc_descriptions(soc_desc_file)
    soc_index_terms = read_soc_coding_index(soc_index_file)
    
    # Categorize all terms
    categorized = defaultdict(set)
    
    for term in naics_terms:
        cat = categorize_term(term, "naics industry")
        categorized[cat].add(term)
    
    for term in soc_desc_terms:
        cat = categorize_term(term, "soc occupation description")
        categorized[cat].add(term)
    
    for term in soc_index_terms:
        cat = categorize_term(term, "soc occupation coding index")
        categorized[cat].add(term)
    
    # Output results
    output = {}
    for cat in sorted(categorized.keys()):
        sorted_terms = sorted(categorized[cat], key=str.lower)
        output[cat] = sorted_terms
        print(f"\n{'='*80}")
        print(f"CATEGORY: {cat} ({len(sorted_terms)} terms)")
        print(f"{'='*80}")
        for t in sorted_terms:
            print(f"  {t}")
    
    # Save as JSON
    output_path = r"C:\careertrojan\data\ai_data_final\gazetteers\classification_collocations.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\n\nSaved {sum(len(v) for v in output.values())} total terms to {output_path}")
    
    # Print summary
    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}")
    for cat in sorted(output.keys()):
        print(f"  {cat}: {len(output[cat])} terms")
    print(f"  TOTAL: {sum(len(v) for v in output.values())} terms")

if __name__ == "__main__":
    main()
