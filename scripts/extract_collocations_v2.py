"""
Extract multi-word collocations from NAICS and SOC classification Excel files.
Phase 1: Read and dump raw terms from each file.
"""
import openpyxl
import re
import json
import sys
from collections import defaultdict

def clean_text(text):
    if not text:
        return ""
    text = str(text).strip()
    text = re.sub(r'^[\s\-–—,;:]+|[\s\-–—,;:]+$', '', text)
    return text

def is_valid_phrase(text):
    """Check if text is a valid multi-word phrase."""
    if not text or not isinstance(text, str):
        return False
    words = text.split()
    if len(words) < 2 or len(words) > 10:
        return False
    if len(text) < 5:
        return False
    if re.match(r'^[\d\s\.\-/]+$', text):
        return False
    # Skip if mostly numbers
    alpha_chars = sum(1 for c in text if c.isalpha())
    if alpha_chars < len(text) * 0.4:
        return False
    return True

def process_naics(filepath):
    print(f"Reading NAICS: {filepath}", flush=True)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    terms = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name}", flush=True)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            for cell in row:
                if cell and isinstance(cell, str) and len(cell) > 4:
                    # Direct cell value
                    cleaned = clean_text(cell)
                    if is_valid_phrase(cleaned):
                        terms.add(cleaned)
                    
                    # Split on semicolons, commas in parentheses, etc.
                    for part in re.split(r'[;\(\)]', cell):
                        part = clean_text(part)
                        if is_valid_phrase(part):
                            terms.add(part)
        print(f"    Processed {row_count} rows", flush=True)
    
    wb.close()
    print(f"  Total NAICS terms: {len(terms)}", flush=True)
    return terms

def process_soc_descriptions(filepath):
    print(f"Reading SOC descriptions: {filepath}", flush=True)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    terms = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name}", flush=True)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            for cell in row:
                if cell and isinstance(cell, str) and len(cell) > 4:
                    cleaned = clean_text(cell)
                    if is_valid_phrase(cleaned):
                        terms.add(cleaned)
                    for part in re.split(r'[;\(\)]', cell):
                        part = clean_text(part)
                        if is_valid_phrase(part):
                            terms.add(part)
        print(f"    Processed {row_count} rows", flush=True)
    
    wb.close()
    print(f"  Total SOC desc terms: {len(terms)}", flush=True)
    return terms

def process_soc_index(filepath):
    print(f"Reading SOC coding index: {filepath}", flush=True)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    terms = set()
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ('INFO', 'FILE SPEC'):
            continue
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name}", flush=True)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            for cell in row:
                if cell and isinstance(cell, str) and len(cell) > 4:
                    cleaned = clean_text(cell)
                    if is_valid_phrase(cleaned):
                        terms.add(cleaned)
            if row_count % 5000 == 0:
                print(f"    ...{row_count} rows", flush=True)
        print(f"    Processed {row_count} rows", flush=True)
    
    wb.close()
    print(f"  Total SOC index terms: {len(terms)}", flush=True)
    return terms

def categorize_term(term, source=""):
    term_lower = term.lower()
    ctx = source.lower()
    combined = term_lower + " " + ctx
    
    oil_gas = ['oil ', 'gas ', 'petroleum', 'drilling', 'pipeline', 'refiner', 'crude',
               'wellhead', 'fracking', 'natural gas', 'offshore', 'subsea', 'petrochemical',
               'fuel ', 'energy', 'mining', 'coal ', 'geothermal', 'solar', 'wind energy',
               'renewable', 'nuclear', 'power plant', 'electric util', 'turbine', 'reservoir',
               'extraction', 'well drilling', 'oil and gas', 'power generat', 'hydroelectric',
               'fossil fuel', 'lng ', 'derrick', 'rig ', ' rig', 'refinery']
    if any(kw in combined for kw in oil_gas):
        return 'OIL_GAS'
    
    bio = ['pharma', 'biotech', 'clinical', 'medical', 'health', 'hospital', 'therapeutic',
           'diagnostic', 'surgical', 'dental', 'nursing', 'physician', 'doctor', 'patient',
           'biolog', 'genetic', 'vaccine', 'drug ', 'medication', 'laboratory', 'patholog',
           'radiolog', 'oncolog', 'cardio', 'neurol', 'optic', 'veterinar', 'epidemiolog',
           'anatomy', 'physiolog', 'immunolog', 'microbiol', 'biochem', 'biomedi',
           'life science', 'healthcare', 'therapist', 'optometr', 'chiropract',
           'psycholog', 'psychiatr', 'ambulance', 'paramedic', 'midwi', 'prosth',
           'orthop', 'audiolog', 'speech therap', 'occupational therap', 'physiother',
           'denti', 'hygieni', 'clinic', 'dispens']
    if any(kw in combined for kw in bio):
        return 'BIOTECH_PHARMA'
    
    fin = ['financ', 'bank', 'insurance', 'invest', 'securit', 'credit', 'loan',
           'mortgage', 'accounting', 'audit', 'tax ', 'fiscal', 'treasury', 'asset manage',
           'wealth', 'broker', 'underwrite', 'actuar', 'hedge fund', 'mutual fund',
           'pension', 'fiduciar', 'compliance', 'risk manage', 'capital market',
           'equity', 'bond ', 'commodit', 'forex', 'stock ', 'portfolio',
           'financial', 'payroll', 'bookkeep', 'chartered account', 'debt', 'revenue',
           'monetary', 'exchang']
    if any(kw in combined for kw in fin):
        return 'FINANCIAL_SERVICES'
    
    tech = ['software', 'computer', 'data ', 'algorithm', 'programm', 'coding',
            'developer', 'information technology', 'cyber', 'network', 'database',
            'cloud', 'machine learn', 'artificial intel', 'web develop', 'system admin',
            'devops', 'full stack', 'front end', 'back end', 'mobile app',
            'user interface', 'user experience', 'information system', 'digital',
            'automation', 'robotics', 'telecomm', 'semiconductor', 'microprocessor',
            'circuit', 'embedded system', 'firmware', 'it support', 'tech support',
            'help desk', 'sas ', 'python', 'java', ' sql', 'electronic', 'satellite',
            'broadband', 'wireless', 'fibre optic', 'fiber optic', 'internet',
            'web design', 'graphic design', 'multimedia', 'video game', 'app develop',
            'it consult', 'it manag', 'it project', 'it direct', 'it special',
            'it engineer', 'computing', 'informatic']
    if any(kw in combined for kw in tech):
        return 'TECH_SKILL'
    
    mfg = ['manufactur', 'assembl', 'production', 'factory', 'industrial',
           'machining', 'welding', 'fabricat', 'forging', 'stamping', 'casting',
           'molding', 'extrusion', 'cnc', 'quality control', 'supply chain',
           'warehouse', 'logistics', 'lean manufactur', 'six sigma', 'injection mold',
           'sheet metal', 'tool and die', 'process engineer', 'plant manager',
           'material handl', 'packaging', 'textile', 'steel', 'aluminum', 'plastic',
           'rubber', 'ceramic', 'glass', 'lumber', 'wood ', 'paper ', 'printing',
           'chemical manufactur', 'auto part', 'aerospace', 'metal work', 'fitter',
           'turner', 'miller', 'lathe', 'press operator', 'machine operator',
           'quality inspect', 'quality assur', 'material test', 'welder',
           'sheet metal worker', 'boilermaker', 'toolmaker']
    if any(kw in combined for kw in mfg):
        return 'MANUFACTURING'
    
    # Default based on source
    if 'soc' in ctx or 'occupation' in ctx or 'coding index' in ctx:
        return 'JOB_TITLE'
    return 'INDUSTRY_TERM'

def main():
    base = r"L:\antigravity_version_ai_data_final\automated_parser"
    
    # Process files
    naics_terms = process_naics(f"{base}\\2022_NAICS_Descriptions.xlsx")
    soc_desc_terms = process_soc_descriptions(f"{base}\\extendedsoc2020structureanddescriptionsexcel03122025.xlsx")
    soc_index_terms = process_soc_index(f"{base}\\soc2020volume2thecodingindexexcel03122025.xlsx")
    
    # Categorize
    categorized = defaultdict(list)
    
    for term in naics_terms:
        cat = categorize_term(term, "naics industry")
        categorized[cat].append(term)
    
    for term in soc_desc_terms:
        cat = categorize_term(term, "soc occupation description")
        categorized[cat].append(term)
    
    for term in soc_index_terms:
        cat = categorize_term(term, "soc occupation coding index")
        categorized[cat].append(term)
    
    # Deduplicate and sort
    output = {}
    for cat in sorted(categorized.keys()):
        unique = sorted(set(categorized[cat]), key=str.lower)
        output[cat] = unique
    
    # Save JSON
    output_path = r"C:\careertrojan\data\ai_data_final\gazetteers\classification_collocations.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    # Print summary
    print(f"\n{'='*80}", flush=True)
    print("SUMMARY", flush=True)
    print(f"{'='*80}", flush=True)
    total = 0
    for cat in sorted(output.keys()):
        count = len(output[cat])
        total += count
        print(f"  {cat}: {count} terms", flush=True)
        # Print first 10 as samples
        for t in output[cat][:10]:
            print(f"    - {t}", flush=True)
        print(f"    ...", flush=True)
    print(f"\n  TOTAL: {total} terms", flush=True)
    print(f"  Saved to: {output_path}", flush=True)

if __name__ == "__main__":
    main()
