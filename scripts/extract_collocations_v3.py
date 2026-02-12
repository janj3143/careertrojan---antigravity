"""
Extract clean multi-word collocations from NAICS and SOC classification Excel files.
Produces high-quality gazetteer entries grouped by category.
V3: Heavy filtering and quality checks.
"""
import openpyxl
import re
import json
from collections import defaultdict

# ── Noise filters ──────────────────────────────────────────────────────────────
NOISE_PATTERNS = [
    r'<[^>]+>',           # HTML tags
    r'[\'\"]\s*>',        # escaped HTML
    r'^\d{4}/\d{2}',      # SOC code references like "2254/02"
    r'^\.[\s\S]*',         # lines starting with .
    r'Illustrative Example',
    r'Cross-Reference',
    r'See also',
    r'added \'',
    r'^\d{4}\s+Framework', # header text
    r'^\d{4}\s+Structure',
    r'Background check',
    r'driving licence',
    r'DBS check',
    r'Manual',
    r'^\(in combination\)',
    r'Establishments in this industry',
    r'This industry comprises',
    r'This u\.s\. industry',
    r'This subsector comprises',
    r'This sector comprises',
    r'This group comprises',
    r'This major group',
    r'industries not covered',
    r'in the Detailed',
]
NOISE_RE = re.compile('|'.join(NOISE_PATTERNS), re.IGNORECASE)

def is_noise(text):
    if NOISE_RE.search(text):
        return True
    # Too short or too generic
    if len(text) < 6:
        return True
    # Mostly non-alpha
    alpha = sum(1 for c in text if c.isalpha())
    if alpha < len(text) * 0.5:
        return True
    # Single-word check 
    if len(text.split()) < 2:
        return True
    return False

def clean(text):
    if not text or not isinstance(text, str):
        return ""
    t = str(text).strip()
    t = re.sub(r'<[^>]+>', ' ', t)       # strip HTML
    t = re.sub(r"['\"]>", '', t)
    t = re.sub(r'\s+', ' ', t).strip()
    t = re.sub(r'^[\s\-–—,;:\.]+|[\s\-–—,;:\.]+$', '', t)
    # Title case normalization for display
    return t

def extract_phrases(cell_text):
    """Return all valid multi-word phrases from a cell."""
    if not cell_text or not isinstance(cell_text, str):
        return []
    phrases = []
    # Split on semicolons and parentheses
    parts = re.split(r'[;\(\)\[\]]', cell_text)
    for part in parts:
        c = clean(part)
        words = c.split()
        if 2 <= len(words) <= 8 and not is_noise(c):
            phrases.append(c)
    return phrases

# ── Categorization ─────────────────────────────────────────────────────────────

def categorize(term, source_hint=""):
    tl = term.lower()

    # OIL_GAS: Energy, petroleum, mining, utilities
    if re.search(r'\b(oil|gas|petroleum|drilling|pipeline|refin|crude|wellhead|frack|'
                 r'natural gas|offshore|subsea|petrochemical|fuel|energy|mining|coal|'
                 r'geothermal|solar|wind\s+(energy|power|farm|turbine)|renewable|'
                 r'nuclear|power\s+(plant|generation|station)|electric\s+util|turbine|'
                 r'reservoir|hydroelectric|fossil\s+fuel|derrick|lng|well\s+drill|'
                 r'oil\s+and\s+gas|extraction|quarr|mineral|ore\b|smelting|peat)\b', tl):
        return 'OIL_GAS'

    # BIOTECH_PHARMA
    if re.search(r'\b(pharma|biotech|clinical|medical|health|hospital|therapeutic|'
                 r'diagnostic|surgical|dental|nursing|physician|doctor|patient|biolog|'
                 r'genetic|vaccine|drug|medication|laboratory|patholog|radiolog|oncolog|'
                 r'cardio|neurol|optic|veterinar|epidemiolog|anatomy|physiolog|immunolog|'
                 r'microbiol|biochem|biomedi|life\s+science|healthcare|therapist|'
                 r'optometr|chiropract|psycholog|psychiatr|ambulance|paramedic|midwif|'
                 r'prosth|orthop|audiolog|speech\s+therap|occupational\s+therap|'
                 r'physiother|denti|hygien|dispensing|surgeon|practitioner|radiat|'
                 r'sonograph|anaesth|anesth|nurse|nurs|ophthalm|podiatr|dietit|'
                 r'medicin|anatomy|clinical\s+trial|bioprocess)\b', tl):
        return 'BIOTECH_PHARMA'

    # FINANCIAL_SERVICES
    if re.search(r'\b(financ|banking|insurance|invest|securities|credit|loan|mortgage|'
                 r'accounting|audit|tax\s|fiscal|treasury|asset\s+manag|wealth|broker|'
                 r'underwrite|actuar|hedge\s+fund|mutual\s+fund|pension|fiduciar|'
                 r'compliance|risk\s+manag|capital\s+market|equity|bond\s|commodit|'
                 r'forex|stock\s|portfolio|payroll|bookkeep|chartered\s+account|'
                 r'debt|revenue|monetary|exchange|clearing\s*house|savings\s+and\s+loan|'
                 r'savings\s+bank|trust\s+office|consumer\s+financ|financial\s+plan|'
                 r'venture\s+capital|private\s+equity|reinsur|claims\s+adjust|'
                 r'loss\s+adjust|claims\s+handler)\b', tl):
        return 'FINANCIAL_SERVICES'

    # TECH_SKILL  — be specific, avoid catching "technician" generically
    if re.search(r'\b(software|computer\s+\w+|data\s+(scien|analy|engineer|architect|'
                 r'warehous|mining|model)|artificial\s+intellig|machine\s+learn|'
                 r'deep\s+learn|cyber\s+secur|cloud\s+comput|web\s+develop|'
                 r'system\s+admini|devops|full\s+stack|front[\s-]end|back[\s-]end|'
                 r'mobile\s+app|user\s+(interface|experience)|information\s+(system|'
                 r'technol)|digital\s+transform|automation\s+engineer|robotics|'
                 r'telecomm|semiconductor|microprocess|embedded\s+system|firmware|'
                 r'it\s+support|tech\s+support|help\s+desk|database\s+admin|'
                 r'network\s+(engineer|admin|architect|secur)|cloud\s+(architect|'
                 r'engineer)|agile\s+|scrum|devops|block\s*chain|internet\s+of\s+things|'
                 r'iot\b|big\s+data|business\s+intellig|erp\b|crm\b|saas\b|'
                 r'cyber|information\s+secur|penetration\s+test|app\s+develop|'
                 r'programming|electronic\s+(engineer|design|manufactur)|'
                 r'circuit\s+design|satellite\s+commun|broadband|wireless\s+commun|'
                 r'fibre\s+optic|fiber\s+optic|graphic\s+design|multimedia|'
                 r'video\s+game|game\s+develop|ux\s+design|ui\s+design|'
                 r'computer\s+science|computing)\b', tl):
        return 'TECH_SKILL'

    # MANUFACTURING
    if re.search(r'\b(manufactur|assembl|production\s+(line|manager|engineer|worker|'
                 r'supervisor)|factory|industrial\s+(engineer|design|mechanic)|'
                 r'machin(ing|ist)|welding|welder|fabricat|forging|stamping|casting|'
                 r'mold(ing|er)|extrusion|cnc|quality\s+(control|assur|inspect)|'
                 r'supply\s+chain|warehou|logistics|lean\s+manufactur|six\s+sigma|'
                 r'injection\s+mold|sheet\s+metal|tool\s+and\s+die|'
                 r'process\s+engineer|plant\s+manager|material\s+handl|packag|'
                 r'textile|steel\s+\w+|aluminum\s+\w+|plastic\s+\w+|printing\s+\w+|'
                 r'chemical\s+manufactur|aerospace|metal\s+work|boilermaker|'
                 r'toolmaker|press\s+operato|machine\s+operato|'
                 r'auto(motive)?\s+(part|manufactur|assembly)|powder\s+coat|'
                 r'surface\s+treat|heat\s+treat|galvaniz|anodiz|plating)\b', tl):
        return 'MANUFACTURING'

    # Fallback
    if 'soc' in source_hint.lower() or 'coding' in source_hint.lower():
        return 'JOB_TITLE'
    return 'INDUSTRY_TERM'


# ── Excel readers ──────────────────────────────────────────────────────────────

def read_naics(fp):
    print(f"[NAICS] {fp}", flush=True)
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    terms = set()
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).lower() if c else '' for c in row]
            print(f"  Headers: {headers}", flush=True)
            continue
        # Row structure: Code | Title | Description
        vals = list(row)
        # Get title column (usually col 1 or 2)
        for ci, cell in enumerate(vals):
            if cell and isinstance(cell, str):
                for phrase in extract_phrases(cell):
                    terms.add(phrase)
    wb.close()
    print(f"  >> {len(terms)} NAICS terms", flush=True)
    return terms

def read_soc_desc(fp):
    print(f"[SOC-DESC] {fp}", flush=True)
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    terms = set()
    for sname in wb.sheetnames:
        if sname == 'INFO' or 'OLD' in sname:
            continue
        ws = wb[sname]
        print(f"  Sheet: {sname}", flush=True)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            for cell in row:
                if cell and isinstance(cell, str):
                    for phrase in extract_phrases(cell):
                        terms.add(phrase)
    wb.close()
    print(f"  >> {len(terms)} SOC description terms", flush=True)
    return terms

def read_soc_index(fp):
    print(f"[SOC-INDEX] {fp}", flush=True)
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    terms = set()
    for sname in wb.sheetnames:
        if sname in ('INFO', 'FILE SPEC', 'Abbreviations', 'ISCO-08 framework'):
            continue
        ws = wb[sname]
        print(f"  Sheet: {sname}", flush=True)
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            for cell in row:
                if cell and isinstance(cell, str):
                    for phrase in extract_phrases(cell):
                        terms.add(phrase)
            count += 1
            if count % 10000 == 0:
                print(f"    ...{count} rows", flush=True)
        print(f"    {count} rows done", flush=True)
    wb.close()
    print(f"  >> {len(terms)} SOC index terms", flush=True)
    return terms


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    base = r"L:\antigravity_version_ai_data_final\automated_parser"

    naics = read_naics(f"{base}\\2022_NAICS_Descriptions.xlsx")
    soc_d = read_soc_desc(f"{base}\\extendedsoc2020structureanddescriptionsexcel03122025.xlsx")
    soc_i = read_soc_index(f"{base}\\soc2020volume2thecodingindexexcel03122025.xlsx")

    cats = defaultdict(set)
    for t in naics:
        cats[categorize(t, 'naics')].add(t)
    for t in soc_d:
        cats[categorize(t, 'soc_desc')].add(t)
    for t in soc_i:
        cats[categorize(t, 'soc_coding_index')].add(t)

    # Build final output
    output = {}
    for cat in sorted(cats.keys()):
        items = sorted(cats[cat], key=str.lower)
        output[cat] = items
    
    # Save
    out_path = r"C:\careertrojan\data\ai_data_final\gazetteers\classification_collocations.json"
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # Summary with samples
    print(f"\n{'='*80}", flush=True)
    print("EXTRACTION COMPLETE", flush=True)
    print(f"{'='*80}", flush=True)
    total = 0
    for cat in sorted(output.keys()):
        n = len(output[cat])
        total += n
        print(f"\n  {cat}: {n} terms", flush=True)
        # Show first 25
        for t in output[cat][:25]:
            print(f"    • {t}", flush=True)
        if n > 25:
            print(f"    ... ({n - 25} more)", flush=True)
    print(f"\n  GRAND TOTAL: {total} unique terms", flush=True)
    print(f"  Output: {out_path}", flush=True)


if __name__ == "__main__":
    main()
